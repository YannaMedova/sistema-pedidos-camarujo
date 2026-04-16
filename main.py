import sys
import os
import json
import pandas as pd
import traceback
from PyQt6.QtGui import QBrush, QColor, QAction, QFont, QPixmap
from datetime import datetime
from openpyxl.drawing.image import Image as ExcelImage
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QPushButton, QTableView,
                             QHeaderView, QFileDialog, QSpinBox, QStyledItemDelegate,
                             QMessageBox, QDialog, QFormLayout, QDoubleSpinBox,
                             QGraphicsOpacityEffect)
from PyQt6.QtCore import Qt, QAbstractTableModel, QSortFilterProxyModel, pyqtSignal, QTimer, QPropertyAnimation, QRect, QEasingCurve
from PyQt6.QtWidgets import QLineEdit
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PyQt6.QtWidgets import QWidget, QPushButton, QLabel, QHBoxLayout
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D

# --- CONFIGURAÇÃO DE TEMAS ---
THEMES = {
    "Claro": """
        QMainWindow { background-color: #F4F7F9; }
        QLabel { color: #102A43; font-family: 'Segoe UI', Arial; }
        QTableView { 
            background-color: #FFFFFF; 
            alternate-background-color: #EBF3FA;
            color: #102A43; 
            gridline-color: #BCCCDC; 
            selection-background-color: #00509E;
            selection-color: white;
            border: 1px solid #BCCCDC;
            border-radius: 4px;
        }
        QHeaderView::section { 
            background-color: #00509E; 
            color: #FFFFFF; 
            padding: 6px; 
            border: 1px solid #003B73; 
            font-weight: bold; 
        }
        QPushButton { background-color: #00509E; color: white; border-radius: 5px; padding: 8px 12px; font-weight: bold; }
        QPushButton:hover { background-color: #003B73; }
    """,
    "Escuro": """
        QMainWindow { background-color: #0D1B2A; }
        QLabel { color: #E0E1DD; font-family: 'Segoe UI', Arial; }
        QTableView { 
            background-color: #1B263B; 
            alternate-background-color: #24354D; 
            color: #E0E1DD; 
            gridline-color: #415A77; 
            selection-background-color: #577590;
            selection-color: white;
            border: 1px solid #415A77;
            border-radius: 4px;
        }
        QHeaderView::section { 
            background-color: #0B2545; 
            color: #E0E1DD; 
            padding: 6px; 
            border: 1px solid #133A66; 
            font-weight: bold; 
        }
        QPushButton { background-color: #133A66; color: white; border-radius: 5px; padding: 8px 12px; font-weight: bold; }
        QPushButton:hover { background-color: #1F528A; }
    """
}


def carregar_config():
    if os.path.exists('config_camarujo.json'):
        with open('config_camarujo.json', 'r') as f:
            return json.load(f)
    return {"tema": "Claro", "margem": 50.0}


def salvar_config(config):
    with open('config_camarujo.json', 'w') as f:
        json.dump(config, f)


ARQUIVO_ESTADO = "ultimo_pedido_camarujo.json"
LOGO_ARQUIVO = "logo_camarujo.png"


def caminho_recurso(nome_arquivo):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, nome_arquivo)
class SplashCamarujo(QWidget):
    def __init__(self):
        super().__init__()

        self.setFixedSize(520, 320)
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)

        self.fundo = QWidget(self)
        self.fundo.setGeometry(0, 0, 520, 320)
        self.fundo.setStyleSheet("""
            background-color: #0B2545;
            border-radius: 18px;
        """)

        self.logo_label = QLabel(self.fundo)
        self.logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        caminho_logo = caminho_recurso(LOGO_ARQUIVO)
        if os.path.exists(caminho_logo):
            logo = QPixmap(caminho_logo)
            logo = logo.scaled(
                220, 220,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            )
            self.logo_label.setPixmap(logo)
        else:
            self.logo_label.setText("🐌")
            self.logo_label.setStyleSheet("font-size: 72px; color: white;")

        # posição inicial da logo (menor)
        self.logo_label.setGeometry(175, 30, 170, 120)

        self.titulo_label = QLabel("CAMARUJO AQUARISMO", self.fundo)
        self.titulo_label.setGeometry(60, 190, 400, 34)
        self.titulo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.titulo_label.setStyleSheet("""
            color: white;
            font-size: 22px;
            font-weight: bold;
            background: transparent;
        """)

        self.subtitulo_label = QLabel("Sistema de Pedidos", self.fundo)
        self.subtitulo_label.setGeometry(60, 225, 400, 26)
        self.subtitulo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.subtitulo_label.setStyleSheet("""
            color: #D9E6F2;
            font-size: 14px;
            background: transparent;
        """)

        self.status_label = QLabel("Carregando...", self.fundo)
        self.status_label.setGeometry(60, 265, 400, 22)
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet("""
            color: #BFD7EA;
            font-size: 12px;
            background: transparent;
        """)

        # efeito de transparência da logo
        self.logo_opacity = QGraphicsOpacityEffect()
        self.logo_label.setGraphicsEffect(self.logo_opacity)
        self.logo_opacity.setOpacity(0.0)

        # animação de fade
        self.anim_fade = QPropertyAnimation(self.logo_opacity, b"opacity")
        self.anim_fade.setDuration(900)
        self.anim_fade.setStartValue(0.0)
        self.anim_fade.setEndValue(1.0)
        self.anim_fade.setEasingCurve(QEasingCurve.Type.InOutCubic)

        # animação de zoom suave
        self.anim_zoom = QPropertyAnimation(self.logo_label, b"geometry")
        self.anim_zoom.setDuration(1100)
        self.anim_zoom.setStartValue(QRect(180, 40, 180, 120))
        self.anim_zoom.setEndValue(QRect(140, 20, 240, 180))
        self.anim_zoom.setEasingCurve(QEasingCurve.Type.OutCubic)

    def iniciar_animacao(self):
        self.anim_fade.start()
        self.anim_zoom.start()

# --- DELEGADOS (WIDGETS DA TABELA) ---
class QuantidadeWidget(QWidget):
    valorAlterado = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.valor = 0
        self.setFixedHeight(28)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(4)

        self.btn_menos = QPushButton("➖")
        self.input_valor = QLineEdit("0")
        self.btn_mais = QPushButton("➕")

        for btn in (self.btn_menos, self.btn_mais):
            btn.setFixedSize(22, 22)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #1F528A;
                    color: white;
                    border: none;
                    border-radius: 11px;
                    font-weight: 900;
                    font-size: 14px;
                    padding: 0px;
                    margin: 0px;
                    text-align: center;
                }
                QPushButton:hover {
                    background-color: #3A7BC8;
                }
                QPushButton:pressed {
                    background-color: #163E69;
                }
            """)

        self.input_valor.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.input_valor.setFixedWidth(38)
        self.input_valor.setStyleSheet("""
            QLineEdit {
                color: #102A43;
                background: rgba(255, 255, 255, 0.85);
                border: 1px solid #7FA6D6;
                border-radius: 6px;
                font-weight: bold;
                padding: 1px 4px;
            }
        """)

        layout.addWidget(self.btn_menos)
        layout.addWidget(self.input_valor)
        layout.addWidget(self.btn_mais)

        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setStyleSheet("background: rgba(0, 0, 0, 0);")

        self.btn_menos.clicked.connect(self.diminuir)
        self.btn_mais.clicked.connect(self.aumentar)
        self.input_valor.editingFinished.connect(self.valor_digitado)
        self.input_valor.returnPressed.connect(self.valor_digitado)

    def setValor(self, valor):
        try:
            self.valor = max(0, int(float(valor)))
        except Exception:
            self.valor = 0
        self.input_valor.setText(str(self.valor))

    def getValor(self):
        return self.valor

    def aumentar(self):
        self.valor += 1
        self.input_valor.setText(str(self.valor))
        self.valorAlterado.emit()

    def diminuir(self):
        if self.valor > 0:
            self.valor -= 1
            self.input_valor.setText(str(self.valor))
            self.valorAlterado.emit()

    def valor_digitado(self):
        texto = self.input_valor.text().strip()

        try:
            valor = int(texto) if texto else 0
            if valor < 0:
                valor = 0
        except Exception:
            valor = self.valor

        self.valor = valor
        self.input_valor.setText(str(self.valor))
        self.valorAlterado.emit()


class QuantityDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QuantidadeWidget(parent)
        editor.valorAlterado.connect(lambda: self.commitData.emit(editor))
        return editor

    def paint(self, painter, option, index):
        super().paint(painter, option, index)

    def setEditorData(self, editor, index):
        source_index = index.model().mapToSource(index)
        valor = source_index.model()._data.iloc[source_index.row(), source_index.column()]
        editor.setValor(valor)

    def setModelData(self, editor, model, index):
        source_index = model.mapToSource(index)
        model.sourceModel().setData(
            source_index,
            editor.getValor(),
            Qt.ItemDataRole.EditRole
        )

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class PriceDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QDoubleSpinBox(parent)
        editor.setMinimum(0)
        editor.setMaximum(99999)
        editor.setDecimals(2)
        editor.setStyleSheet("QDoubleSpinBox { background: white; color: black; border: 1px solid #ccc; }")
        editor.valueChanged.connect(lambda: self.commitData.emit(editor))
        return editor

    def setEditorData(self, editor, index):
        source_index = index.model().mapToSource(index)
        valor = source_index.model()._data.iloc[source_index.row(), source_index.column()]
        try:
            editor.setValue(float(valor))
        except:
            editor.setValue(0.0)

    def setModelData(self, editor, model, index):
        source_index = model.mapToSource(index)
        model.sourceModel().setData(source_index, editor.value(), Qt.ItemDataRole.EditRole)


class CamarujoModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data
        self.headers = [
            "Pedido", "Qtd Min", "Item", "Tamanho",
            "Custo (R$)", "Sugerido (R$)",
            "Venda Manual (R$)", "Valor Total Bruto (R$)", "Valor Custo Total (R$)"
        ]

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        return len(self.headers)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None

        coluna = index.column()
        valor = self._data.iloc[index.row(), coluna]

        if role == Qt.ItemDataRole.BackgroundRole:
            try:
                pedido = self._data.iloc[index.row(), 0]
                if int(float(pedido)) > 0:
                    return QBrush(QColor("#1E3A5F"))
            except:
                pass

        if role in (Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.EditRole):
            if pd.isna(valor):
                return ""

            if coluna in [4, 5, 6, 7, 8]:
                if role == Qt.ItemDataRole.DisplayRole:
                    return self.formatar_moeda_br(valor)
                else:
                    try:
                        return f"{float(valor):.2f}"
                    except:
                        return "0.00"

            if coluna in [0, 1]:
                try:
                    if float(valor) == 0 and coluna == 1:
                        return "-"
                    return str(int(float(valor)))
                except:
                    return str(valor)

            return str(valor)

        if role == Qt.ItemDataRole.TextAlignmentRole:
            if coluna in [0, 1, 3, 4, 5, 6, 7, 8]:
                return Qt.AlignmentFlag.AlignCenter

        return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if role != Qt.ItemDataRole.EditRole or not index.isValid():
            return False

        linha = index.row()
        coluna = index.column()

        try:
            if coluna == 0:
                self._data.iloc[linha, coluna] = int(value)
            elif coluna == 6:
                if value in ("", None):
                    self._data.iloc[linha, coluna] = 0.0
                else:
                    self._data.iloc[linha, coluna] = float(value)
            else:
                return False

            pedido = self._data.iloc[linha, 0]
            custo = self._data.iloc[linha, 4]
            sugerido = self._data.iloc[linha, 5]
            manual = self._data.iloc[linha, 6]

            pedido = 0 if pd.isna(pedido) else int(pedido)
            custo = 0.0 if pd.isna(custo) else float(custo)
            sugerido = 0.0 if pd.isna(sugerido) else float(sugerido)
            manual = 0.0 if pd.isna(manual) else float(manual)

            valor_unitario_venda = manual if manual > 0 else sugerido

            self._data.iloc[linha, 7] = valor_unitario_venda * pedido
            self._data.iloc[linha, 8] = custo * pedido

            self.dataChanged.emit(self.index(linha, 0), self.index(linha, 8))
            return True

        except Exception:
            print(traceback.format_exc())
            return False

    def flags(self, index):
        if index.column() in [0, 6]:
            return Qt.ItemFlag.ItemIsEditable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
        return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable

    def headerData(self, section, orientation, role):
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self.headers[section]
        return None

    def formatar_moeda_br(self, valor):
        try:
            valor = float(valor)
            texto = f"{valor:,.2f}"
            texto = texto.replace(",", "X").replace(".", ",").replace("X", ".")
            return f"R$ {texto}"
        except:
            return "R$ 0,00"


class ConfigDialog(QDialog):
    def __init__(self, parent, margem_atual):
        super().__init__(parent)
        self.setWindowTitle("Configurações")
        self.setFixedSize(300, 150)
        layout = QFormLayout(self)
        self.margem_spin = QDoubleSpinBox()
        self.margem_spin.setRange(0, 500)
        self.margem_spin.setValue(margem_atual)
        layout.addRow("Margem Padrão (%):", self.margem_spin)
        btn_salvar = QPushButton("Salvar Alterações")
        btn_salvar.clicked.connect(self.accept)
        layout.addWidget(btn_salvar)


class FiltroPeixeProxyModel(QSortFilterProxyModel):
    def __init__(self):
        super().__init__()
        self.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)

    def filterAcceptsRow(self, source_row, source_parent):
        if not self.filterRegularExpression().pattern():
            return True

        model = self.sourceModel()
        item_index = model.index(source_row, 2, source_parent)
        tamanho_index = model.index(source_row, 3, source_parent)

        item = str(model.data(item_index, Qt.ItemDataRole.DisplayRole) or "")
        tamanho = str(model.data(tamanho_index, Qt.ItemDataRole.DisplayRole) or "")

        texto = f"{item} {tamanho}"
        return self.filterRegularExpression().match(texto).hasMatch()


class JanelaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = carregar_config()
        self.df = pd.DataFrame(columns=[
            "Pedido", "Qtd Min", "Item", "Tamanho",
            "Custo", "Sugerido", "Venda Manual", "Valor Total Bruto", "Valor Custo Total"
        ])
        self.initUI()
        self.aplicar_estilo()
        self.carregar_ultimo_pedido()

    def initUI(self):
        self.setWindowTitle("SISTEMA DE PEDIDOS CAMARUJO AQUARISMO")
        self.resize(1200, 750)
        layout_principal = QVBoxLayout()

        topo = QHBoxLayout()

        self.lbl_logo = QLabel()
        caminho_logo = caminho_recurso(LOGO_ARQUIVO)

        if os.path.exists(caminho_logo):
            pixmap = QPixmap(caminho_logo)
            pixmap = pixmap.scaled(
                42, 42,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            )
            self.lbl_logo.setPixmap(pixmap)
        else:
            self.lbl_logo.setText("🐌")
            self.lbl_logo.setStyleSheet("font-size: 26px;")

        self.lbl_titulo = QLabel("SISTEMA DE PEDIDOS - CAMARUJO AQUARISMO")
        self.lbl_titulo.setStyleSheet("font-size: 20px; font-weight: bold;")

        btn_importar = QPushButton("📥 Importar Planilha")
        btn_importar.clicked.connect(self.importar_excel)

        btn_tema = QPushButton("🌙 / ☀️")
        btn_tema.setToolTip("Alternar tema claro/escuro")
        btn_tema.clicked.connect(self.alternar_tema)

        btn_hist = QPushButton("📅 Histórico")
        btn_hist.clicked.connect(self.abrir_historico)

        btn_config = QPushButton("⚙️ Config")
        btn_config.clicked.connect(self.abrir_config)

        btn_novo = QPushButton("🆕 Novo Pedido")
        btn_novo.clicked.connect(self.novo_pedido)

        self.input_filtro = QLineEdit()
        self.input_filtro.setPlaceholderText("Buscar peixe...")
        self.input_filtro.setFixedWidth(260)

        icone_lupa = self.style().standardIcon(self.style().StandardPixmap.SP_FileDialogContentsView)
        action_lupa = QAction(icone_lupa, "", self.input_filtro)
        self.input_filtro.addAction(action_lupa, QLineEdit.ActionPosition.LeadingPosition)

        self.input_filtro.setStyleSheet("""
            QLineEdit {
                border-radius: 6px;
                border: 1px solid #ccc;
                height: 28px;
                padding-left: 5px;
            }
        """)

        topo.addWidget(self.lbl_logo)
        topo.addWidget(self.lbl_titulo)
        topo.addStretch()
        topo.addWidget(self.input_filtro)
        topo.addSpacing(10)
        topo.addWidget(btn_importar)
        topo.addWidget(btn_tema)
        topo.addWidget(btn_hist)
        topo.addWidget(btn_config)
        topo.addWidget(btn_novo)

        self.tabela = QTableView()
        self.tabela.setAlternatingRowColors(True)
        self.model = CamarujoModel(self.df)

        self.proxy_model = FiltroPeixeProxyModel()
        self.proxy_model.setSourceModel(self.model)

        self.tabela.setModel(self.proxy_model)
        self.input_filtro.textChanged.connect(self.proxy_model.setFilterFixedString)
        self.input_filtro.textChanged.connect(self.reabrir_editores_pedido)

        self.delegate_qty = QuantityDelegate(self.tabela)
        self.delegate_preco = PriceDelegate(self.tabela)
        self.tabela.setItemDelegateForColumn(0, self.delegate_qty)
        self.tabela.setItemDelegateForColumn(6, self.delegate_preco)

        self.tabela.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabela.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.tabela.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.tabela.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)

        self.tabela.setColumnWidth(0, 150)
        self.tabela.setColumnWidth(1, 70)
        self.tabela.setColumnWidth(3, 80)
        self.tabela.setColumnWidth(6, 120)
        self.tabela.setColumnWidth(7, 140)
        self.tabela.setColumnWidth(8, 140)

        rodape = QHBoxLayout()
        self.lbl_resumo = QLabel("Total Custo: R$ 0,00  |  Lucro Estimado: R$ 0,00")
        self.lbl_resumo.setStyleSheet(
            "font-size: 15px; font-weight: bold; border-top: 1px solid #ccc; padding-top: 10px;"
        )

        btn_exp_cliente = QPushButton("📄 Exportar p/ Cliente")
        btn_exp_cliente.clicked.connect(self.exportar_cliente)

        btn_exp_fornecedor = QPushButton("📦 Fechar Pedido Fornecedor")
        btn_exp_fornecedor.clicked.connect(self.exportar_fornecedor)

        rodape.addWidget(self.lbl_resumo)
        rodape.addStretch()
        rodape.addWidget(btn_exp_cliente)
        rodape.addWidget(btn_exp_fornecedor)

        layout_principal.addLayout(topo)
        layout_principal.addWidget(self.tabela)
        layout_principal.addSpacing(10)
        layout_principal.addLayout(rodape)

        container = QWidget()
        container.setLayout(layout_principal)
        self.setCentralWidget(container)

    def reabrir_editores_pedido(self):
        for row in range(self.proxy_model.rowCount()):
            self.tabela.openPersistentEditor(self.proxy_model.index(row, 0))

    def aplicar_dataframe_na_tabela(self):
        self.model = CamarujoModel(self.df)
        self.proxy_model.setSourceModel(self.model)
        self.model.dataChanged.connect(self.atualizar_totais)
        self.model.dataChanged.connect(self.salvar_estado_atual)
        self.atualizar_totais()
        self.reabrir_editores_pedido()

    def salvar_estado_atual(self):
        try:
            if self.df is None or self.df.empty:
                return

            df_salvar = self.df.copy()

            for coluna in ['Pedido', 'Qtd Min']:
                if coluna in df_salvar.columns:
                    df_salvar[coluna] = pd.to_numeric(df_salvar[coluna], errors='coerce').fillna(0).astype(int)

            for coluna in ['Custo', 'Sugerido', 'Venda Manual', 'Valor Total Bruto', 'Valor Custo Total']:
                if coluna in df_salvar.columns:
                    df_salvar[coluna] = pd.to_numeric(df_salvar[coluna], errors='coerce').fillna(0.0)

            with open(ARQUIVO_ESTADO, "w", encoding="utf-8") as f:
                json.dump(df_salvar.to_dict(orient="records"), f, ensure_ascii=False, indent=2)

        except Exception:
            print(traceback.format_exc())

    def carregar_ultimo_pedido(self):
        try:
            if not os.path.exists(ARQUIVO_ESTADO):
                return

            with open(ARQUIVO_ESTADO, "r", encoding="utf-8") as f:
                dados = json.load(f)

            if not dados:
                return

            self.df = pd.DataFrame(dados)

            colunas_esperadas = [
                "Pedido", "Qtd Min", "Item", "Tamanho",
                "Custo", "Sugerido", "Venda Manual", "Valor Total Bruto", "Valor Custo Total"
            ]

            for coluna in colunas_esperadas:
                if coluna not in self.df.columns:
                    if coluna in ["Pedido", "Qtd Min"]:
                        self.df[coluna] = 0
                    elif coluna in ["Custo", "Sugerido", "Venda Manual", "Valor Total Bruto", "Valor Custo Total"]:
                        self.df[coluna] = 0.0
                    else:
                        self.df[coluna] = ""

            self.df = self.df[colunas_esperadas]

            self.df["Pedido"] = pd.to_numeric(self.df["Pedido"], errors="coerce").fillna(0).astype(int)
            self.df["Qtd Min"] = pd.to_numeric(self.df["Qtd Min"], errors="coerce").fillna(0).astype(int)

            for coluna in ["Custo", "Sugerido", "Venda Manual", "Valor Total Bruto", "Valor Custo Total"]:
                self.df[coluna] = pd.to_numeric(self.df[coluna], errors="coerce").fillna(0.0)

            self.aplicar_dataframe_na_tabela()

        except Exception:
            print(traceback.format_exc())

    def novo_pedido(self):
        resposta = QMessageBox.question(
            self,
            "Novo Pedido",
            "Deseja iniciar um novo pedido?\n\nO pedido atual salvo será substituído.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if resposta != QMessageBox.StandardButton.Yes:
            return

        self.df = pd.DataFrame(columns=[
            "Pedido", "Qtd Min", "Item", "Tamanho",
            "Custo", "Sugerido", "Venda Manual", "Valor Total Bruto", "Valor Custo Total"
        ])

        if os.path.exists(ARQUIVO_ESTADO):
            os.remove(ARQUIVO_ESTADO)

        self.aplicar_dataframe_na_tabela()
        self.importar_excel()

    def importar_excel(self):
        caminho, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Planilhas (*.xlsx *.xls *.csv)")
        if not caminho:
            return

        try:
            extensao = os.path.splitext(caminho)[1].lower()

            if extensao in ['.xlsx', '.xls']:
                temp_df = pd.read_excel(caminho, header=None)
            elif extensao == '.csv':
                try:
                    temp_df = pd.read_csv(caminho, header=None, sep=';', encoding='utf-8-sig')
                    if len(temp_df.columns) == 1:
                        temp_df = pd.read_csv(caminho, header=None, sep=',', encoding='utf-8-sig')
                except Exception:
                    temp_df = pd.read_csv(caminho, header=None, sep=',', encoding='latin1')
            else:
                QMessageBox.warning(self, "Formato Inválido", "Selecione um arquivo .xlsx, .xls ou .csv.")
                return

            if temp_df is None or temp_df.empty:
                QMessageBox.warning(self, "Aviso", "A planilha parece estar vazia ou não pôde ser lida!")
                return

            start_row = 0
            for i in range(min(20, len(temp_df))):
                linha_texto = ' '.join(map(str, temp_df.iloc[i].fillna(''))).lower()
                if 'descrição' in linha_texto or 'descricao' in linha_texto or 'valor' in linha_texto:
                    start_row = i + 1
                    break

            temp_df = temp_df.iloc[start_row:].reset_index(drop=True)

            while len(temp_df.columns) < 6:
                temp_df[len(temp_df.columns)] = ""

            temp_df = temp_df.dropna(subset=[1])
            temp_df = temp_df[temp_df.iloc[:, 1].astype(str).str.strip() != '']
            temp_df = temp_df[temp_df.iloc[:, 1].astype(str).str.strip().str.lower() != 'nan']

            self.df = pd.DataFrame()
            self.df['Pedido'] = [0] * len(temp_df)
            self.df['Qtd Min'] = pd.to_numeric(temp_df.iloc[:, 0], errors='coerce').fillna(0).astype(int)
            self.df['Item'] = temp_df.iloc[:, 1].fillna('').astype(str).str.strip()
            self.df['Tamanho'] = temp_df.iloc[:, 2].fillna('').astype(str).str.strip().replace('nan', '')
            self.df['Custo'] = pd.to_numeric(temp_df.iloc[:, 3], errors='coerce').fillna(0.0)

            margem = float(self.config.get('margem', 50.0))
            self.df['Sugerido'] = (self.df['Custo'] * (1 + margem / 100)).round(2)
            self.df['Venda Manual'] = 0.0
            self.df['Valor Total Bruto'] = 0.0
            self.df['Valor Custo Total'] = 0.0

            self.aplicar_dataframe_na_tabela()
            self.salvar_estado_atual()

        except Exception:
            erro_completo = traceback.format_exc()
            QMessageBox.critical(self, "Erro Detalhado",
                                 f"Houve uma falha.\nPor favor, me mostre este texto:\n\n{erro_completo}")

    def atualizar_totais(self):
        valor_custo_total = pd.to_numeric(self.df['Valor Custo Total'], errors='coerce').fillna(0)
        valor_total_bruto = pd.to_numeric(self.df['Valor Total Bruto'], errors='coerce').fillna(0)

        total_custo = valor_custo_total.sum()
        lucro = valor_total_bruto.sum() - total_custo

        self.lbl_resumo.setText(
            f"Total Custo: {self.model.formatar_moeda_br(total_custo)}  |  "
            f"Lucro Estimado: {self.model.formatar_moeda_br(lucro)}"
        )

    def recalcular_tabela_com_margem(self):
        if self.df is None or self.df.empty:
            return

        margem = float(self.config.get('margem', 50.0))

        self.df['Sugerido'] = (
            pd.to_numeric(self.df['Custo'], errors='coerce').fillna(0.0) * (1 + margem / 100)
        ).round(2)

        pedido = pd.to_numeric(self.df['Pedido'], errors='coerce').fillna(0).astype(int)
        custo = pd.to_numeric(self.df['Custo'], errors='coerce').fillna(0.0)
        sugerido = pd.to_numeric(self.df['Sugerido'], errors='coerce').fillna(0.0)
        manual = pd.to_numeric(self.df['Venda Manual'], errors='coerce').fillna(0.0)

        valor_unitario_venda = manual.where(manual > 0, sugerido)

        self.df['Valor Total Bruto'] = (valor_unitario_venda * pedido).round(2)
        self.df['Valor Custo Total'] = (custo * pedido).round(2)

        self.aplicar_dataframe_na_tabela()
        self.salvar_estado_atual()

    def exportar_cliente(self):
        if self.df.empty:
            return

        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Tabela Cliente",
            f"Precos_Camarujo_{datetime.now().strftime('%d_%m')}.xlsx",
            "Excel (*.xlsx)"
        )

        if not caminho:
            return

        try:
            df_cliente = self.df[['Item', 'Tamanho', 'Sugerido', 'Venda Manual']].copy()

            df_cliente['Preço de Venda'] = df_cliente.apply(
                lambda row: row['Venda Manual'] if float(row['Venda Manual']) > 0 else row['Sugerido'],
                axis=1
            )

            df_cliente = df_cliente[['Item', 'Tamanho', 'Preço de Venda']]
            df_cliente = df_cliente[df_cliente['Item'].astype(str).str.strip() != '']

            df_cliente.to_excel(caminho, index=False, startrow=5)

            wb = load_workbook(caminho)
            ws = wb.active
            ws.title = "Tabela de Preços"

            caminho_logo = caminho_recurso(LOGO_ARQUIVO)

            # faixa superior
            ws.merge_cells("A1:C2")
            for row in range(1, 3):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="0B2545")

            # logo
            if os.path.exists(caminho_logo):
                img = ExcelImage(caminho_logo)
                img.width = 190
                img.height = 220

                marker = AnchorMarker(
                    col=0,
                    colOff=1050000,
                    row=0,
                    rowOff=20000
                )

                img.anchor = OneCellAnchor(
                    _from=marker,
                    ext=XDRPositiveSize2D(2400000, 1400000)
                )

                ws.add_image(img)

            # título
            ws.merge_cells("A3:C3")
            ws["A3"] = "CAMARUJO AQUARISMO"
            ws["A3"].font = Font(color="FFFFFF", bold=True, size=16)
            ws["A3"].fill = PatternFill(fill_type="solid", fgColor="133A66")
            ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

            # subtítulo
            ws.merge_cells("A4:C4")
            ws["A4"] = f"Tabela de Preços - {datetime.now().strftime('%d/%m/%Y')}"
            ws["A4"].font = Font(color="0B2545", bold=True, size=11)
            ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A4"].fill = PatternFill(fill_type="solid", fgColor="F4F7F9")

            fill_cabecalho = PatternFill(fill_type="solid", fgColor="133A66")
            fill_linha_clara = PatternFill(fill_type="solid", fgColor="F8FBFF")
            fill_linha_escura = PatternFill(fill_type="solid", fgColor="EEF4FA")
            fonte_branca = Font(color="FFFFFF", bold=True)

            borda_fina = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )

            # cabeçalho da tabela = linha 6
            for cell in ws[6]:
                cell.fill = fill_cabecalho
                cell.font = fonte_branca
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borda_fina

            # conteúdo começa na linha 7
            for row_num in range(7, ws.max_row + 1):
                fill_atual = fill_linha_clara if row_num % 2 == 0 else fill_linha_escura

                for col_num in range(1, 4):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = borda_fina
                    cell.fill = fill_atual
                    cell.font = Font(color="102A43", size=11)

            # alinhamentos
            for row in range(7, ws.max_row + 1):
                ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
                ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws[f"C{row}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

            # moeda
            for row in range(7, ws.max_row + 1):
                ws[f"C{row}"].number_format = 'R$ #,##0.00'

            # larguras
            ws.column_dimensions["A"].width = 36
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 18

            # alturas
            ws.row_dimensions[1].height = 80
            ws.row_dimensions[3].height = 24
            ws.row_dimensions[4].height = 20
            ws.row_dimensions[6].height = 22

            for row in range(7, ws.max_row + 1):
                ws.row_dimensions[row].height = 20

            ws.freeze_panes = "A7"

            wb.save(caminho)

            QMessageBox.information(self, "Sucesso", "Tabela para clientes exportada com sucesso!")

        except Exception:
            erro_completo = traceback.format_exc()
            QMessageBox.critical(
                self,
                "Erro ao Exportar",
                f"Houve uma falha ao exportar a tabela:\n\n{erro_completo}"
            )

    def exportar_fornecedor(self):
        pedido = self.df[self.df['Pedido'] > 0].copy()
        if pedido.empty:
            QMessageBox.warning(self, "Aviso", "Adicione quantidade a pelo menos um item para fechar o pedido.")
            return

        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Pedido Fornecedor",
            f"Pedido_Camarujo_{datetime.now().strftime('%d_%m')}.xlsx",
            "Excel (*.xlsx)"
        )
        if not caminho:
            return

        try:
            df_fornecedor = pedido[
                ['Item', 'Tamanho', 'Pedido', 'Custo', 'Valor Custo Total']
            ].copy()

            df_fornecedor.columns = [
                'Item',
                'Tamanho',
                'Quantidade',
                'Custo Unitário (R$)',
                'Total (R$)'
            ]

            total_pedido = pd.to_numeric(
                df_fornecedor['Total (R$)'], errors='coerce'
            ).fillna(0).sum()

            linha_total = pd.DataFrame([{
                'Item': '',
                'Tamanho': '',
                'Quantidade': '',
                'Custo Unitário (R$)': 'TOTAL DO PEDIDO',
                'Total (R$)': total_pedido
            }])

            df_exportacao = pd.concat([df_fornecedor, linha_total], ignore_index=True)

            df_exportacao.to_excel(caminho, index=False, startrow=5)

            wb = load_workbook(caminho)
            ws = wb.active
            ws.title = "Pedido Fornecedor"

            caminho_logo = caminho_recurso(LOGO_ARQUIVO)

            # faixa superior
            ws.merge_cells("A1:E2")
            for row in range(1, 3):
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="0B2545")

            # logo
            if os.path.exists(caminho_logo):
                img = ExcelImage(caminho_logo)
                img.width = 190
                img.height = 220

                marker = AnchorMarker(
                    col=0,
                    colOff=1950000,
                    row=0,
                    rowOff=20000
                )

                img.anchor = OneCellAnchor(
                    _from=marker,
                    ext=XDRPositiveSize2D(2400000, 1400000)
                )

                ws.add_image(img)

            # título
            ws.merge_cells("A3:E3")
            ws["A3"] = "CAMARUJO AQUARISMO"
            ws["A3"].font = Font(color="FFFFFF", bold=True, size=16)
            ws["A3"].fill = PatternFill(fill_type="solid", fgColor="133A66")
            ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

            # subtítulo
            ws.merge_cells("A4:E4")
            ws["A4"] = f"Pedido ao Fornecedor - {datetime.now().strftime('%d/%m/%Y')}"
            ws["A4"].font = Font(color="0B2545", bold=True, size=11)
            ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A4"].fill = PatternFill(fill_type="solid", fgColor="F4F7F9")

            fill_cabecalho = PatternFill(fill_type="solid", fgColor="133A66")
            fill_linha_clara = PatternFill(fill_type="solid", fgColor="F8FBFF")
            fill_linha_escura = PatternFill(fill_type="solid", fgColor="EEF4FA")
            fill_total = PatternFill(fill_type="solid", fgColor="0B2545")

            fonte_branca = Font(color="FFFFFF", bold=True)
            fonte_total = Font(color="FFFFFF", bold=True, size=12)

            borda_fina = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )

            # cabeçalho da tabela = linha 6
            for cell in ws[6]:
                cell.fill = fill_cabecalho
                cell.font = fonte_branca
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borda_fina

            linha_total_excel = ws.max_row

            # conteúdo começa na linha 7
            for row_num in range(7, ws.max_row + 1):
                is_total = (row_num == linha_total_excel)
                fill_atual = fill_total if is_total else (
                    fill_linha_clara if row_num % 2 == 0 else fill_linha_escura
                )

                for col_num in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = borda_fina
                    cell.fill = fill_atual

                    if is_total:
                        cell.font = fonte_total
                    else:
                        cell.font = Font(color="102A43", size=11)

            # alinhamentos
            for row in range(7, ws.max_row + 1):
                ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
                ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
                ws[f"E{row}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

            # moeda
            for row in range(7, ws.max_row + 1):
                if row != linha_total_excel:
                    ws[f"D{row}"].number_format = 'R$ #,##0.00'
                ws[f"E{row}"].number_format = 'R$ #,##0.00'

            # larguras
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 18

            # alturas
            ws.row_dimensions[1].height = 80
            ws.row_dimensions[3].height = 24
            ws.row_dimensions[4].height = 20
            ws.row_dimensions[6].height = 22

            for row in range(7, ws.max_row + 1):
                ws.row_dimensions[row].height = 20

            ws.freeze_panes = "A7"

            wb.save(caminho)

            if not os.path.exists('historico'):
                os.makedirs('historico')

            pedido.to_csv(
                f"historico/pedido_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.csv",
                index=False
            )

            QMessageBox.information(self, "Sucesso", "Pedido fechado e arquivado no histórico!")

        except Exception:
            erro_completo = traceback.format_exc()
            QMessageBox.critical(
                self,
                "Erro ao Exportar",
                f"Houve uma falha ao exportar o pedido:\n\n{erro_completo}"
            )

    def abrir_config(self):
        dialog = ConfigDialog(self, self.config['margem'])
        if dialog.exec():
            self.config['margem'] = dialog.margem_spin.value()
            salvar_config(self.config)
            self.recalcular_tabela_com_margem()

            QMessageBox.information(
                self,
                "Configurações",
                "Margem atualizada e aplicada automaticamente à tabela atual!"
            )

    def abrir_historico(self):
        if not os.path.exists('historico'):
            os.makedirs('historico')
        os.startfile('historico') if sys.platform == 'win32' else os.system('open historico')

    def alternar_tema(self):
        self.config['tema'] = "Claro" if self.config['tema'] == "Escuro" else "Escuro"
        self.aplicar_estilo()
        salvar_config(self.config)

    def aplicar_estilo(self):
        estilo_base = THEMES[self.config['tema']]
        estilo_extra = """
            QTableView::item {
                background: transparent;
            }

            QMessageBox {
                background-color: #1E1E1E;
            }

            QMessageBox QLabel {
                color: white;
                font-size: 13px;
            }

            QMessageBox QPushButton {
                background-color: #2E6EB5;
                color: white;
                padding: 6px 12px;
                border-radius: 6px;
            }

            QMessageBox QPushButton:hover {
                background-color: #3F83D1;
            }
        """
        self.setStyleSheet(estilo_base + estilo_extra)

    def closeEvent(self, event):
        self.salvar_estado_atual()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    splash = SplashCamarujo()
    splash.show()
    splash.iniciar_animacao()
    app.processEvents()

    win = JanelaPrincipal()

    def mostrar_janela():
        win.show()
        splash.close()

    QTimer.singleShot(2200, mostrar_janela)

    sys.exit(app.exec())